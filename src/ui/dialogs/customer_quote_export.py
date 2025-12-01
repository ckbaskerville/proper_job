"""Customer quote export functionality."""

import tkinter as tk
from tkinter import messagebox
from typing import Any, List

from src.business.calculator import QuoteCalculator
from src.models import Project


class CustomerQuoteExporter:
    """Handles exporting customer quotes to Excel."""

    def __init__(
        self,
        calculator: QuoteCalculator,
        project: Project,
        fitting_cost: float = 0.0,
        extras_cost: float = 0.0
    ):
        """Initialize the customer quote exporter.

        Args:
            calculator: Quote calculator with optimization results
            project: Current project
            fitting_cost: Fitting cost to include in quote
            extras_cost: Extras cost to include in quote
        """
        self.calculator = calculator
        self.project = project
        self.fitting_cost = fitting_cost
        self.extras_cost = extras_cost

    def export_customer_quote(self) -> None:
        """Export customer quote to Excel with unit and component pricing."""
        from tkinter import filedialog

        filename = filedialog.asksaveasfilename(
            title="Export Customer Quote (Excel)",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

        if not filename:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        except ImportError:
            messagebox.showerror(
                "Missing Dependency",
                "openpyxl is required to export Excel files.\n"
                "Please install it with 'pip install openpyxl'."
            )
            return

        try:
            # Get unit breakdown from calculator
            breakdowns = self.calculator.calculate_unit_breakdown()
            
            if not breakdowns:
                messagebox.showinfo(
                    "No Data",
                    "There is no quote data to export."
                )
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Customer Quote"

            # Define styles
            header_fill = PatternFill(start_color="FFB3D9FF", end_color="FFB3D9FF", fill_type="solid")  # Light blue
            unit_fill = PatternFill(start_color="FFB3D9FF", end_color="FFB3D9FF", fill_type="solid")  # Light blue
            header_font = Font(bold=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Write header row
            headers = ['Name', 'Component', 'Notes', 'Component Quantity', 'Quantity', 'Unit Cost', 'Total Cost']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            current_row = 2
            total_cost = 0.0

            # Process each unit
            for unit_breakdown in breakdowns:
                # Calculate unit costs with markup
                unit_subtotal = unit_breakdown.unit_subtotal
                markup = self.calculator.labor_manager.calculate_markup(unit_subtotal)
                unit_cost_with_markup = unit_subtotal + markup
                unit_total_cost = unit_cost_with_markup * unit_breakdown.quantity
                total_cost += unit_total_cost

                # Write unit row (light blue background)
                for col_idx in range(1, 8):
                    cell = ws.cell(row=current_row, column=col_idx)
                    if col_idx == 1:
                        cell.value = unit_breakdown.unit_name
                    elif col_idx == 4:
                        # Component Quantity - empty for whole cabinets
                        cell.value = ""
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_idx == 5:
                        # Quantity - number of cabinets
                        cell.value = unit_breakdown.quantity
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_idx == 6:
                        cell.value = round(unit_cost_with_markup, 2)
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_idx == 7:
                        cell.value = round(unit_total_cost, 2)
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.fill = unit_fill
                    cell.border = thin_border
                current_row += 1

                # Write component rows (white background)
                # Get the unit to access component quantities
                unit = self.project.cabinets[unit_breakdown.unit_index]
                
                for component in unit_breakdown.components:
                    # Calculate component costs with markup
                    component_subtotal = component.total_cost
                    component_markup = self.calculator.labor_manager.calculate_markup(component_subtotal)
                    component_cost_with_markup = component_subtotal + component_markup
                    component_total_cost = component_cost_with_markup * unit_breakdown.quantity

                    # Format component name - use "Drawers" instead of "Drawer 1", "Drawer 2", etc.
                    component_name = component.component_name
                    if component_name.startswith("Drawer"):
                        component_name = "Drawers"
                    elif component_name == "Face Frame":
                        component_name = "Faceframe"

                    # Get component quantity
                    component_quantity = self._get_component_quantity(component, unit)

                    # Write all cells in the component row with borders
                    # Column A (Name) - empty for components
                    cell = ws.cell(row=current_row, column=1)
                    cell.border = thin_border
                    
                    # Component name goes in column B
                    cell = ws.cell(row=current_row, column=2, value=component_name)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="top")
                    
                    # Format notes with detailed information
                    notes = self._format_component_notes(component, unit_breakdown)
                    cell = ws.cell(row=current_row, column=3, value=notes)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="top")
                    
                    # Component Quantity
                    cell = ws.cell(row=current_row, column=4, value=component_quantity if component_quantity > 0 else "")
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                    
                    # Quantity - empty for components (only whole cabinets have quantity)
                    cell = ws.cell(row=current_row, column=5, value="")
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                    
                    # Unit Cost
                    cell = ws.cell(row=current_row, column=6, value=round(component_cost_with_markup, 2))
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="right", vertical="top")
                    
                    # Total Cost
                    cell = ws.cell(row=current_row, column=7, value=round(component_total_cost, 2))
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="right", vertical="top")
                    
                    current_row += 1

            # Add rows for Fitting and Extras (same color as header)
            fitting_row_cost = self.fitting_cost
            extras_row_cost = self.extras_cost
            total_cost += fitting_row_cost + extras_row_cost
            
            for row_label, row_cost in [("Fitting", fitting_row_cost), ("Extras", extras_row_cost)]:
                for col_idx in range(1, 8):
                    cell = ws.cell(row=current_row, column=col_idx)
                    if col_idx == 1:
                        cell.value = row_label
                        cell.font = Font(bold=True)
                    elif col_idx == 7:  # Total Cost column
                        cell.value = f"{row_cost:.2f}"
                    cell.fill = header_fill
                    cell.border = thin_border
                    if col_idx == 1:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif col_idx >= 5:  # Quantity, Unit Cost, Total Cost columns
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                current_row += 1

            # Write total row
            ws.cell(row=current_row, column=6, value="Total Cost").font = Font(bold=True)
            ws.cell(row=current_row, column=6).border = thin_border
            ws.cell(row=current_row, column=6).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=current_row, column=7, value=round(total_cost, 2)).font = Font(bold=True)
            ws.cell(row=current_row, column=7).border = thin_border
            ws.cell(row=current_row, column=7).alignment = Alignment(horizontal="right", vertical="center")

            # Adjust column widths
            from openpyxl.utils import get_column_letter
            column_widths = {
                'A': 20,  # Name
                'B': 15,  # Component
                'C': 70,  # Notes (increased for more detail)
                'D': 20,  # Component Quantity
                'E': 10,  # Quantity
                'F': 12,  # Unit Cost
                'G': 12   # Total Cost
            }
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            wb.save(filename)
            messagebox.showinfo(
                "Export Complete",
                f"Customer quote exported to {filename}"
            )

        except Exception as e:
            messagebox.showerror(
                "Export Error",
                f"Failed to export customer quote:\n{str(e)}"
            )

    def _format_component_notes(self, component: Any, unit_breakdown: Any) -> str:
        """Format component notes with detailed information."""
        notes_parts = []
        
        # Get the unit to access component details
        unit = self.project.cabinets[unit_breakdown.unit_index]
        
        if component.component_name == "Carcass":
            # Material and finish
            material_parts = []
            if unit.carcass.sprayed:
                material_parts.append("Sprayed")
            material_parts.append(unit.carcass.material)
            if unit.carcass.material_thickness:
                material_parts.append(f"{unit.carcass.material_thickness}mm")
            notes_parts.append(", ".join(material_parts))
            
            # Dimensions
            dims = f"{unit.carcass.width:.0f} (W) × {unit.carcass.height:.0f} (H) × {unit.carcass.depth:.0f} (D)"
            notes_parts.append(dims)
            
            # Shelves
            if unit.carcass.shelves > 0:
                shelf_text = f"{unit.carcass.shelves} Shelves" if unit.carcass.shelves > 1 else "1 Shelf"
                notes_parts.append(shelf_text)
            
            # Back panel
            if unit.carcass.has_back:
                notes_parts.append("With back panel")
            else:
                notes_parts.append("No back panel")
        
        elif component.component_name.startswith("Drawer"):
            # Get drawer details
            drawer_idx = int(component.component_name.split()[-1]) - 1 if component.component_name.split()[-1].isdigit() else 0
            if drawer_idx < len(unit.drawers):
                drawer = unit.drawers[drawer_idx]
                
                # Material
                drawer_material = f"{drawer.material} {drawer.thickness}mm"
                notes_parts.append(drawer_material)
                
                # Dimensions
                drawer_depth, drawer_width = drawer.calculate_drawer_dimensions()
                drawer_dims = f"{drawer.height:.0f} (H) × {drawer_width:.0f} (W) × {drawer_depth:.0f} (D)"
                notes_parts.append(drawer_dims)
                
                # Runner details
                if drawer.runner_model:
                    runner_info = f"{drawer.runner_model} {drawer.runner_size}mm, {drawer.runner_capacity}kg capacity"
                    notes_parts.append(runner_info)
        
        elif component.component_name == "Doors":
            if unit.doors:
                # Material and finish
                door_material = []
                if unit.doors.sprayed:
                    door_material.append("Sprayed")
                door_material.append(unit.doors.material)
                if unit.doors.material_thickness:
                    door_material.append(f"{unit.doors.material_thickness}mm")
                notes_parts.append(", ".join(door_material))
                
                # Door type
                notes_parts.append(unit.doors.door_type.lower().capitalize())
                
                # Handle
                if unit.doors.cut_handle:
                    notes_parts.append("Cut handle")
                
                # Hinges
                if unit.doors.hinge_type:
                    hinge_info = f"{unit.doors.hinge_type} Hinges"
                    if unit.doors.hinges_per_door > 0:
                        hinge_info += f" ({unit.doors.hinges_per_door} per door)"
                    notes_parts.append(hinge_info)
                
                # Quantity
                if unit.doors.quantity > 0:
                    door_qty = f"{unit.doors.quantity} door{'s' if unit.doors.quantity > 1 else ''}"
                    notes_parts.append(door_qty)
        
        elif component.component_name == "Face Frame":
            if unit.face_frame:
                # Material
                frame_material = f"{unit.face_frame.material} {unit.face_frame.thickness}mm"
                notes_parts.append(frame_material)
                
                # Dimensions
                frame_dims = f"{unit.carcass.height:.0f} (H) × {unit.carcass.width:.0f} (W)"
                notes_parts.append(frame_dims)
                
                # Moulding
                if unit.face_frame.moulding:
                    notes_parts.append("With moulding")
                
                # Finish
                if unit.face_frame.sprayed:
                    notes_parts.append("Sprayed")
        
        return ", ".join(notes_parts)

    def _get_component_quantity(self, component: Any, unit: Any) -> int:
        """Get the quantity for a component.
        
        Args:
            component: ComponentBreakdown object
            unit: Cabinet unit object
            
        Returns:
            Component quantity (e.g., 2 for 2 doors), or 0 if not applicable
        """
        component_name = component.component_name
        
        if component_name == "Doors":
            # Return door quantity
            if unit.doors and unit.doors.quantity > 0:
                return unit.doors.quantity
            return 0
        elif component_name.startswith("Drawer") and "DBC" not in component_name:
            # Each drawer component represents 1 drawer
            return 1
        elif "DBC Drawer" in component_name:
            # Each DBC drawer component represents 1 drawer
            return 1
        elif component_name == "Carcass":
            return 1
        elif component_name == "Face Frame":
            return 1
        else:
            return 0

