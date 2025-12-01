"""Cut list dialog for displaying optimized cutting layouts."""

import tkinter as tk
from tkinter import ttk, messagebox
from typing import Dict, Any, List, Optional, Tuple
from collections import defaultdict

from ..base import BaseDialog
from src.business.calculator import QuoteCalculator
from src.models import Project
import openpyxl


class CutListDialog(BaseDialog):
    """Dialog for displaying cut list information."""

    COMPONENT_ORDER = ["Carcass", "Drawers", "Doors", "Face Frame"]

    def __init__(
            self,
            parent: tk.Widget,
            quote: Dict[str, Any],
            calculator: QuoteCalculator,
            project: Project
    ):
        """Initialize cut list dialog.

        Args:
            parent: Parent widget
            quote: Quote data
            calculator: Quote calculator with optimization results
            project: Current project
        """
        self.quote = quote
        self.calculator = calculator
        self.project = project

        super().__init__(parent, "Cut List", width=900, height=600)

    def _create_body(self, parent: ttk.Frame) -> None:
        """Create dialog body."""
        # Notebook for different views
        notebook = ttk.Notebook(parent)
        notebook.pack(fill=tk.BOTH, expand=True)

        # Summary tab
        summary_frame = ttk.Frame(notebook)
        notebook.add(summary_frame, text="Summary")
        self._create_summary_tab(summary_frame)

        # Detailed cut list tab
        detailed_frame = ttk.Frame(notebook)
        notebook.add(detailed_frame, text="Detailed Cut List")
        self._create_detailed_tab(detailed_frame)

        # By sheet tab
        sheet_frame = ttk.Frame(notebook)
        notebook.add(sheet_frame, text="By Sheet")
        self._create_sheet_tab(sheet_frame)

    def _create_summary_tab(self, parent: ttk.Frame) -> None:
        """Create summary tab."""
        # Title
        ttk.Label(
            parent,
            text="Cut List Summary",
            font=('Arial', 14, 'bold')
        ).pack(pady=10)

        # Summary frame
        summary_frame = ttk.LabelFrame(parent, text="Material Summary", padding=10)
        summary_frame.pack(fill=tk.X, padx=10, pady=5)

        # Create treeview for summary
        columns = ('material', 'thickness', 'sheets', 'parts', 'efficiency', 'cost')
        self.summary_tree = ttk.Treeview(
            summary_frame,
            columns=columns,
            show='headings',
            height=10
        )

        # Configure columns
        self.summary_tree.heading('material', text='Material', anchor='w')
        self.summary_tree.heading('thickness', text='Thickness (mm)', anchor='w')
        self.summary_tree.heading('sheets', text='Sheets', anchor='w')
        self.summary_tree.heading('parts', text='Parts', anchor='w')
        self.summary_tree.heading('efficiency', text='Efficiency', anchor='w')
        self.summary_tree.heading('cost', text='Cost', anchor='w')

        self.summary_tree.column('material', width=150)
        self.summary_tree.column('thickness', width=100)
        self.summary_tree.column('sheets', width=80)
        self.summary_tree.column('parts', width=80)
        self.summary_tree.column('efficiency', width=100)
        self.summary_tree.column('cost', width=100)

        # Scrollbar
        scrollbar = ttk.Scrollbar(
            summary_frame,
            orient=tk.VERTICAL,
            command=self.summary_tree.yview
        )
        self.summary_tree.configure(yscrollcommand=scrollbar.set)

        self.summary_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Populate summary
        self._populate_summary()

        # Totals
        totals_frame = ttk.Frame(parent)
        totals_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Label(
            totals_frame,
            text=f"Total Sheets: {self.quote['total_sheets_required']}",
            font=('Arial', 11, 'bold')
        ).pack(side=tk.LEFT, padx=10)

        ttk.Label(
            totals_frame,
            text=f"Total Material Cost: £{self.quote['material_cost']:.2f}",
            font=('Arial', 11, 'bold')
        ).pack(side=tk.LEFT, padx=10)

    def _populate_summary(self) -> None:
        """Populate summary tree."""
        sheets_breakdown = self.quote.get('sheets_breakdown', {})

        for (material, thickness), data in sheets_breakdown.items():
            efficiency = data.get('efficiency', 0) * 100

            values = (
                material,
                thickness,
                data['sheets_required'],
                data['parts_count'],
                f"{efficiency:.1f}%",
                f"£{data['material_cost']:.2f}"
            )

            self.summary_tree.insert('', tk.END, values=values)

    def _create_detailed_tab(self, parent: ttk.Frame) -> None:
        """Create detailed cut list tab."""
        # Controls
        control_frame = ttk.Frame(parent)
        control_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(control_frame, text="Group by:").pack(side=tk.LEFT, padx=5)

        self.group_var = tk.StringVar(value="Component Type")
        self.group_combo = ttk.Combobox(
            control_frame,
            textvariable=self.group_var,
            values=["Component Type", "Cabinet", "Material", "Size"],
            state="readonly",
            width=15
        )
        self.group_combo.pack(side=tk.LEFT)
        self.group_combo.bind('<<ComboboxSelected>>', self._on_group_changed)

        ttk.Button(
            control_frame,
            text="Export Excel",
            command=self._export_detailed_excel
        ).pack(side=tk.RIGHT, padx=5)

        ttk.Button(
            control_frame,
            text="Export CSV",
            command=self._export_detailed_csv
        ).pack(side=tk.RIGHT, padx=5)

        # Detailed tree - columns will be set dynamically based on grouping
        self.detailed_tree = ttk.Treeview(
            parent,
            columns=(),  # Will be set dynamically
            show='tree headings',
            height=20
        )

        # Scrollbar
        scrollbar = ttk.Scrollbar(
            parent,
            orient=tk.VERTICAL,
            command=self.detailed_tree.yview
        )
        self.detailed_tree.configure(yscrollcommand=scrollbar.set)

        self.detailed_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 10))

        # Populate detailed list
        self._populate_detailed()
        
    def _on_group_changed(self, event=None) -> None:
        """Handle group selection change."""
        self._populate_detailed()

    def _populate_detailed(self) -> None:
        """Populate detailed cut list."""
        # Clear existing
        for item in self.detailed_tree.get_children():
            self.detailed_tree.delete(item)

        group_by = self.group_var.get()
        
        if group_by == "Component Type":
            self._populate_by_component_type()
        elif group_by == "Cabinet":
            self._populate_by_cabinet()
        else:
            # Default to cabinet grouping for other options
            self._populate_by_cabinet()
    
    def _populate_by_component_type(self) -> None:
        """Populate cut list grouped by component type with part columns."""
        all_part_types, grouped_data = self._build_component_grouping_data()

        if not grouped_data:
            return

        columns = ['material', 'quantity', 'cabinets'] + all_part_types
        self.detailed_tree['columns'] = columns

        # Configure columns
        self.detailed_tree.heading('#0', text='Component Type', anchor='w')
        self.detailed_tree.heading('material', text='Material', anchor='w')
        self.detailed_tree.heading('quantity', text='Quantity', anchor='w')
        self.detailed_tree.heading('cabinets', text='Cabinets', anchor='w')

        for part_type in all_part_types:
            self.detailed_tree.heading(part_type, text=part_type, anchor='w')

        # Set column widths
        self.detailed_tree.column('#0', width=120)
        self.detailed_tree.column('material', width=150)
        self.detailed_tree.column('quantity', width=70)
        self.detailed_tree.column('cabinets', width=150)
        for part_type in all_part_types:
            self.detailed_tree.column(part_type, width=120)

        # Populate tree in defined component order
        for comp_type in self.COMPONENT_ORDER:
            rows = grouped_data.get(comp_type)
            if not rows:
                continue

            comp_node = self.detailed_tree.insert(
                '',
                tk.END,
                text=comp_type,
                open=True
            )

            for row_data in rows:
                values = [
                    row_data['material'],
                    row_data['quantity'],
                    ', '.join(sorted(row_data['cabinets']))
                ]

                for part_type in all_part_types:
                    values.append(row_data['parts'].get(part_type, ''))

                self.detailed_tree.insert(comp_node, tk.END, values=values)
    
    def _build_component_grouping_data(self) -> Tuple[List[str], Dict[str, List[Dict[str, Any]]]]:
        """Build data structures for component type grouping."""
        all_part_types = set()
        grouped_data: Dict[str, List[Dict[str, Any]]] = {}

        for comp_type in self.COMPONENT_ORDER:
            signature_map: Dict[Tuple[str, Tuple[Tuple[str, str], ...]], Dict[str, Any]] = {}

            for cabinet in self.project.cabinets:
                parts_dict, material, quantity_increment = self._collect_component_parts(cabinet, comp_type)
                if not material or not parts_dict or quantity_increment == 0:
                    continue

                all_part_types.update(parts_dict.keys())
                signature_key = (material, tuple(sorted(parts_dict.items())))

                if signature_key not in signature_map:
                    signature_map[signature_key] = {
                        'material': material,
                        'quantity': 0,
                        'cabinets': set(),
                        'parts': dict(parts_dict)
                    }

                signature_map[signature_key]['quantity'] += quantity_increment
                signature_map[signature_key]['cabinets'].add(cabinet.carcass.name)

            if signature_map:
                rows = sorted(
                    signature_map.values(),
                    key=lambda row: (row['material'], sorted(row['cabinets']))
                )
                grouped_data[comp_type] = rows

        return sorted(all_part_types), grouped_data

    def _collect_component_parts(
            self,
            cabinet,
            component_type: str
    ) -> Tuple[Dict[str, str], Optional[str], int]:
        """Collect part information for a specific component type."""
        parts: Dict[str, str] = {}
        material: Optional[str] = None
        quantity_multiplier = cabinet.quantity

        def add_part(part_type: str, dims: str) -> None:
            if part_type in parts:
                if dims not in parts[part_type]:
                    parts[part_type] = f"{parts[part_type]}, {dims}"
            else:
                parts[part_type] = dims

        if component_type == "Carcass":
            for part in cabinet.carcass.get_parts():
                part_type = self._get_part_type_name(part.component_type or "carcass", part.id)
                dims = f"{part.width:.0f} (W) × {part.height:.0f} (H)"
                add_part(part_type, dims)
                if not material:
                    material = f"{cabinet.carcass.material} {cabinet.carcass.material_thickness}mm"

        elif component_type == "Drawers":
            if not cabinet.drawers:
                return {}, None, 0
            # Collect parts from all drawers
            for drawer in cabinet.drawers:
                for part in drawer.get_parts():
                    part_type = self._get_part_type_name(part.component_type or "drawer", part.id)
                    dims = f"{part.width:.0f} (W) × {part.height:.0f} (H)"
                    add_part(part_type, dims)
                if not material:
                    material = f"{drawer.material} {drawer.thickness}mm"
            # Quantity is cabinet quantity (all drawers are included in the parts)
            quantity_multiplier = cabinet.quantity

        elif component_type == "Doors":
            if not (cabinet.doors and cabinet.doors.quantity > 0):
                return {}, None, 0
            door_parts = cabinet.doors.get_parts()
            if not door_parts:
                return {}, None, 0
            for part in door_parts:
                part_type = self._get_part_type_name(part.component_type or "door", part.id)
                dims = f"{part.width:.0f} (W) × {part.height:.0f} (H)"
                add_part(part_type, dims)
            material = f"{cabinet.doors.material} {cabinet.doors.material_thickness}mm"
            quantity_multiplier = cabinet.quantity * len(door_parts)

        elif component_type == "Face Frame":
            if not cabinet.face_frame:
                return {}, None, 0
            for part in cabinet.face_frame.get_parts():
                part_type = self._get_part_type_name(part.component_type or "face_frame", part.id)
                dims = f"{part.width:.0f} (W) × {part.height:.0f} (H)"
                add_part(part_type, dims)
            material = f"{cabinet.face_frame.material} {cabinet.face_frame.thickness}mm"

        else:
            return {}, None, 0

        return parts, material, quantity_multiplier

    def _get_part_type_name(self, component_type: Optional[str], part_id: Optional[str] = None) -> str:
        """Convert component_type string to readable part type name.
        
        Args:
            component_type: The component type string
            part_id: Optional part ID to distinguish between similar component types
                    (e.g., top vs bottom rail in face frames)
        """
        if not component_type:
            return 'Unknown'
        
        if component_type.startswith('carcass_'):
            part = component_type.replace('carcass_', '')
            # Handle special cases
            if part == 'back':
                return 'Back'
            elif part == 'side':
                return 'Sides'  # Use plural for Excel column
            elif part == 'top':
                return 'Top'
            elif part == 'bottom':
                return 'Base'  # Merge bottom and base to Base
            elif part == 'shelf':
                return 'Shelves'  # Use plural for Excel column
            return part.capitalize()
        elif component_type.startswith('drawer_'):
            part = component_type.replace('drawer_', '')
            if part == 'front':
                return 'Front'
            elif part == 'back':
                return 'Back'
            elif part == 'side':
                return 'Side'
            elif part == 'base':
                return 'Base'  # Merge bottom and base to Base
            return part.capitalize()
        elif component_type == 'door':
            return 'Door'
        elif component_type.startswith('face_frame_'):
            part = component_type.replace('face_frame_', '')
            if part == 'rail':
                # Use part_id to distinguish top from bottom rail
                if part_id and '_face_frame_top' in part_id:
                    return 'Top'
                elif part_id and '_face_frame_bottom' in part_id:
                    return 'Base'  # Bottom rail becomes Base
                return 'Rail'
            elif part == 'stile':
                return 'Sides'  # Styles become Sides
            return part.capitalize()
        return component_type.capitalize()
    
    def _populate_by_cabinet(self) -> None:
        """Populate detailed cut list grouped by cabinet (original method)."""
        # Set columns for cabinet grouping
        columns = ('id', 'cabinet', 'component', 'material', 'dimensions', 'quantity')
        self.detailed_tree['columns'] = columns
        
        # Configure columns
        self.detailed_tree.heading('#0', text='Group', anchor='w')
        self.detailed_tree.heading('id', text='Part ID', anchor='w')
        self.detailed_tree.heading('cabinet', text='Cabinet', anchor='w')
        self.detailed_tree.heading('component', text='Component', anchor='w')
        self.detailed_tree.heading('material', text='Material', anchor='w')
        self.detailed_tree.heading('dimensions', text='Dimensions (mm)', anchor='w')
        self.detailed_tree.heading('quantity', text='Quantity', anchor='w')

        self.detailed_tree.column('#0', width=150)
        self.detailed_tree.column('id', width=100)
        self.detailed_tree.column('cabinet', width=120)
        self.detailed_tree.column('component', width=100)
        self.detailed_tree.column('material', width=120)
        self.detailed_tree.column('dimensions', width=120)
        self.detailed_tree.column('quantity', width=50)

        # Group parts by cabinet
        for cab_idx, cabinet in enumerate(self.project.cabinets):
            # Create cabinet node
            cab_node = self.detailed_tree.insert(
                '',
                tk.END,
                text=f"{cabinet.carcass.name} (×{cabinet.quantity})",
                open=True
            )

            # Add carcass parts
            carcass_node = self.detailed_tree.insert(
                cab_node,
                tk.END,
                text="Carcass"
            )

            for part in cabinet.carcass.get_parts():
                values = (
                    part.id,
                    cabinet.carcass.name,
                    "Carcass",
                    f"{cabinet.carcass.material} {cabinet.carcass.material_thickness}mm",
                    f"{part.width:.0f} × {part.height:.0f}",
                    cabinet.quantity
                )
                self.detailed_tree.insert(carcass_node, tk.END, values=values)

            # Add drawer parts
            if cabinet.drawers:
                drawers_node = self.detailed_tree.insert(
                    cab_node,
                    tk.END,
                    text="Drawers"
                )

                for drawer in cabinet.drawers:
                    for part in drawer.get_parts():
                        values = (
                            part.id,
                            cabinet.carcass.name,
                            "Drawer",
                            f"{drawer.material} {drawer.thickness}mm",
                            f"{part.width:.0f} × {part.height:.0f}",
                            cabinet.quantity
                        )
                        self.detailed_tree.insert(drawers_node, tk.END, values=values)

            # Add door parts
            if cabinet.doors and cabinet.doors.quantity > 0:
                doors_node = self.detailed_tree.insert(
                    cab_node,
                    tk.END,
                    text="Doors"
                )

                for part in cabinet.doors.get_parts():
                    values = (
                        part.id,
                        cabinet.carcass.name,
                        "Door",
                        f"{cabinet.doors.material} {cabinet.doors.material_thickness}mm",
                        f"{part.width:.0f} × {part.height:.0f}",
                        cabinet.quantity
                    )
                    self.detailed_tree.insert(doors_node, tk.END, values=values)

            # Add face frame parts
            if cabinet.face_frame:
                face_frame_node = self.detailed_tree.insert(
                    cab_node,
                    tk.END,
                    text="Face Frame"
                )

                for part in cabinet.face_frame.get_parts():
                    values = (
                        part.id,
                        cabinet.carcass.name,
                        "Face Frame",
                        f"{cabinet.face_frame.material} {cabinet.face_frame.thickness}mm",
                        f"{part.width:.0f} × {part.height:.0f}",
                        cabinet.quantity
                    )
                    self.detailed_tree.insert(face_frame_node, tk.END, values=values)

    def _create_sheet_tab(self, parent: ttk.Frame) -> None:
        """Create by-sheet tab."""
        # Material selector
        control_frame = ttk.Frame(parent)
        control_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(control_frame, text="Material:").pack(side=tk.LEFT, padx=5)

        self.sheet_material_var = tk.StringVar()
        materials = list(self.calculator._sheets_by_material.keys())

        material_strings = [f"{mat} {thick}mm" for mat, thick in materials]

        self.material_combo = ttk.Combobox(
            control_frame,
            textvariable=self.sheet_material_var,
            values=material_strings,
            state="readonly",
            width=25
        )
        self.material_combo.pack(side=tk.LEFT, padx=5)
        self.material_combo.bind('<<ComboboxSelected>>', self._on_material_selected)

        ttk.Button(
            control_frame,
            text="Export All Sheets",
            command=self._export_all_sheets
        ).pack(side=tk.RIGHT, padx=5)

        # Sheet display
        self.sheet_text = tk.Text(
            parent,
            wrap=tk.NONE,
            font=('Courier', 10),
            height=25
        )
        self.sheet_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Scrollbars
        v_scrollbar = ttk.Scrollbar(
            self.sheet_text,
            orient=tk.VERTICAL,
            command=self.sheet_text.yview
        )
        h_scrollbar = ttk.Scrollbar(
            self.sheet_text,
            orient=tk.HORIZONTAL,
            command=self.sheet_text.xview
        )

        self.sheet_text.configure(
            yscrollcommand=v_scrollbar.set,
            xscrollcommand=h_scrollbar.set
        )

        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

        # Select first material if available
        if material_strings:
            self.sheet_material_var.set(material_strings[0])
            self._on_material_selected()

    def _on_material_selected(self, event=None) -> None:
        """Handle material selection for sheet view."""
        material_str = self.sheet_material_var.get()
        if not material_str:
            return

        # Parse material and thickness
        parts = material_str.rsplit(' ', 1)
        material = parts[0]
        thickness = float(parts[1].replace('mm', ''))

        # Get sheets for this material
        key = (material, thickness)
        sheets = self.calculator._sheets_by_material.get(key, [])

        # Clear text
        self.sheet_text.delete(1.0, tk.END)

        # Display sheets
        for sheet_idx, sheet in enumerate(sheets):
            self.sheet_text.insert(tk.END, f"Sheet {sheet_idx + 1}\n")
            self.sheet_text.insert(tk.END, "=" * 60 + "\n")

            # Calculate efficiency
            sheet_area = self.calculator.sheet_width * self.calculator.sheet_height
            used_area = sum(rect.width * rect.height for rect in sheet)
            efficiency = (used_area / sheet_area) * 100

            self.sheet_text.insert(
                tk.END,
                f"Efficiency: {efficiency:.1f}% | Parts: {len(sheet)}\n\n"
            )

            # List parts
            for rect in sheet:
                self.sheet_text.insert(
                    tk.END,
                    f"  {rect.id}: {rect.width:.0f} × {rect.height:.0f} mm "
                    f"at ({rect.x:.0f}, {rect.y:.0f})\n"
                )

            self.sheet_text.insert(tk.END, "\n")

    def _export_detailed_csv(self) -> None:
        """Export detailed cut list to CSV."""
        from tkinter import filedialog
        import csv

        filename = filedialog.asksaveasfilename(
            title="Export Cut List",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )

        if filename:
            try:
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)

                    # Header
                    writer.writerow([
                        'Part ID', 'Cabinet', 'Component', 'Material',
                        'Width (mm)', 'Height (mm)', 'Quantity'
                    ])

                    # Write parts
                    for cabinet in self.project.cabinets:
                        # Carcass parts
                        for part in cabinet.carcass.get_parts():
                            writer.writerow([
                                part.id,
                                cabinet.carcass.name,
                                'Carcass',
                                f"{cabinet.carcass.material} {cabinet.carcass.material_thickness}mm",
                                part.width,
                                part.height,
                                cabinet.quantity
                            ])

                        # Drawer parts
                        for drawer in cabinet.drawers:
                            for part in drawer.get_parts():
                                writer.writerow([
                                    part.id,
                                    cabinet.carcass.name,
                                    'Drawer',
                                    f"{drawer.material} {drawer.thickness}mm",
                                    part.width,
                                    part.height,
                                    cabinet.quantity
                                ])

                        # Door parts
                        for part in cabinet.doors.get_parts():
                            writer.writerow([
                                part.id,
                                cabinet.carcass.name,
                                'Door',
                                f"{cabinet.doors.material} {cabinet.doors.material_thickness}mm",
                                part.width,
                                part.height,
                                cabinet.quantity
                            ])

                        for part in cabinet.face_frame.get_parts():
                            writer.writerow([
                                part.id,
                                cabinet.carcass.name,
                                'Face Frame',
                                f"{cabinet.face_frame.material} {cabinet.face_frame.thickness}mm",
                                part.width,
                                part.height,
                                cabinet.quantity
                            ])

                messagebox.showinfo(
                    "Export Complete",
                    f"Cut list exported to {filename}"
                )

            except Exception as e:
                messagebox.showerror(
                    "Export Error",
                    f"Failed to export cut list: {str(e)}"
                )

    def _export_detailed_excel(self) -> None:
        """Export detailed cut list to an Excel file with component-specific columns."""
        from tkinter import filedialog

        filename = filedialog.asksaveasfilename(
            title="Export Cut List (Excel)",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

        if not filename:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        except ImportError:
            messagebox.showerror(
                "Missing Dependency",
                "openpyxl is required to export Excel files.\n"
                "Please install it with 'pip install openpyxl'."
            )
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Cut List"

            # Color scheme for component types
            color_map = {
                "Carcass": "FFE3F2FD",      # Light blue
                "Drawers": "FFFDEBD0",      # Light orange
                "Doors": "FFF9E79F",        # Light yellow
                "Face Frame": "FFE8F5E9"    # Light green
            }
            header_fill = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
            header_font = Font(bold=True)
            
            # Define border style for grid
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            current_row = 1

            # Process each component type
            for comp_type in self.COMPONENT_ORDER:
                rows_data = self._get_component_rows_for_excel(comp_type)
                if not rows_data:
                    continue

                # Get column headers for this component type
                headers = self._get_excel_headers_for_component(comp_type)
                
                # Get color for this component type
                row_fill_color = color_map.get(comp_type, "FFFFFFFF")
                component_fill = PatternFill(start_color=row_fill_color, end_color=row_fill_color, fill_type="solid")
                
                # Write component type header with color across all columns
                from openpyxl.utils import get_column_letter
                num_cols = len(headers)
                
                # Write component name to first cell with formatting
                component_cell = ws.cell(row=current_row, column=1, value=comp_type)
                component_cell.font = Font(bold=True, size=12)
                component_cell.fill = component_fill
                component_cell.alignment = Alignment(horizontal="left", vertical="center")
                component_cell.border = thin_border
                
                # Fill remaining cells with the same color and border
                for col_idx in range(2, num_cols + 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.fill = component_fill
                    cell.border = thin_border
                
                # Merge cells for the component header across all columns
                if num_cols > 1:
                    start_col = get_column_letter(1)
                    end_col = get_column_letter(num_cols)
                    merge_range = f'{start_col}{current_row}:{end_col}{current_row}'
                    ws.merge_cells(merge_range)
                    # Re-apply alignment to the merged cell using the first cell reference
                    component_cell.alignment = Alignment(horizontal="left", vertical="center")
                
                current_row += 1

                # Write column headers
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                
                # Ensure all header columns have borders (in case headers list is shorter than expected)
                for col_idx in range(len(headers) + 1, num_cols + 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.border = thin_border
                    cell.fill = header_fill

                current_row += 1

                # Write data rows (no background color)
                for row_data in rows_data:
                    values = row_data['values']
                    # Ensure we have values for all columns (pad with empty strings if needed)
                    num_headers = len(headers)
                    while len(values) < num_headers:
                        values.append('')
                    
                    for col_idx, value in enumerate(values[:num_headers], start=1):
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.alignment = Alignment(horizontal="center", vertical="top")
                        cell.border = thin_border
                    
                    # Add borders to any remaining empty columns
                    for col_idx in range(len(values) + 1, num_headers + 1):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.border = thin_border

                    current_row += 1

                # Add blank row between sections
                current_row += 1

            # Adjust column widths
            from openpyxl.utils import get_column_letter
            for col_idx, column_cells in enumerate(ws.columns, start=1):
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                adjusted_width = min(max(max_length + 2, 12), 60)
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = adjusted_width

            wb.save(filename)
            messagebox.showinfo(
                "Export Complete",
                f"Detailed cut list exported to {filename}"
            )

        except Exception as e:
            messagebox.showerror(
                "Export Error",
                f"Failed to export Excel cut list:\n{str(e)}"
            )

    def _get_excel_headers_for_component(self, component_type: str) -> List[str]:
        """Get column headers for a specific component type."""
        if component_type == "Carcass":
            return ['Material', 'Quantity', 'Unit', 'Unit Size', 'Base', 'Sides', 'Back', 'Top', 'Dividers', 'Shelves', "Cut", "Edged", "Lamello", "Primer", "Lacquered", "Assembled"]
        elif component_type == "Drawers":
            return ['Material', 'Quantity', 'Unit', 'Unit Size', 'Base', 'Back', 'Side', 'Front', 'Cut and Planed', 'Edged', 'Lacquered', 'Sanded', 'Assembled', 'Oiled']
        elif component_type == "Doors":
            return ['Material', 'Quantity', 'Unit', 'Unit Size', 'Dimensions', 'Cut', 'Assembled', 'Primed', 'Top Coat']
        elif component_type == "Face Frame":
            return ['Material', 'Quantity', 'Unit', 'Unit Size', 'Base', 'Sides', 'Top', 'Cut', 'Edged', 'Lamello', 'Primer', 'Lacquered', 'Assembled']
        else:
            return ['Material', 'Quantity', 'Unit', 'Unit Size']

    def _get_component_rows_for_excel(self, component_type: str) -> List[Dict[str, Any]]:
        """Get rows for a component type without grouping - one row per cabinet."""
        rows = []

        for cabinet in self.project.cabinets:
            parts_dict, material, quantity = self._collect_component_parts(cabinet, component_type)
            if not material or not parts_dict or quantity == 0:
                continue

            # Get unit name and unit size (cabinet dimensions)
            unit_name = cabinet.carcass.name
            unit_size = f"{cabinet.carcass.width:.0f} (W) x {cabinet.carcass.height:.0f} (H) x {cabinet.carcass.depth:.0f} (D)"

            # Build values list based on component type
            headers = self._get_excel_headers_for_component(component_type)
            values = [material, quantity, unit_name, unit_size]

            if component_type == "Carcass":
                values.extend([
                    parts_dict.get('Base', ''),
                    parts_dict.get('Sides', ''),
                    parts_dict.get('Back', ''),
                    parts_dict.get('Top', ''),
                    parts_dict.get('Dividers', ''),
                    parts_dict.get('Shelves', '')
                ])
            elif component_type == "Drawers":
                values.extend([
                    parts_dict.get('Base', ''),
                    parts_dict.get('Back', ''),
                    parts_dict.get('Side', '') or parts_dict.get('Sides', ''),
                    parts_dict.get('Front', ''),  # Front column
                    '',  # Cut and Planed (checkbox column)
                ])
            elif component_type == "Doors":
                # For doors, "Dimensions" contains the door dimensions, "Cut" is for checkboxes
                cut_dims = []
                for part_type, dims in parts_dict.items():
                    cut_dims.append(dims)
                
                values.extend([
                    ', '.join(cut_dims) if cut_dims else '',  # Dimensions column with door dimensions
                    ''  # Cut column (checkbox column)
                ])
            elif component_type == "Face Frame":
                values.extend([
                    parts_dict.get('Base', ''),
                    parts_dict.get('Sides', ''),
                    parts_dict.get('Top', ''),
                ])

            rows.append({
                'cabinet_name': cabinet.carcass.name,
                'values': values
            })

        return rows

    def _export_all_sheets(self) -> None:
        """Export all sheet layouts."""
        from tkinter import filedialog

        directory = filedialog.askdirectory(
            title="Select Export Directory"
        )

        if directory:
            try:
                import os

                for (material, thickness), sheets in self.calculator._sheets_by_material.items():
                    # Create file for each material
                    filename = f"{material}_{thickness}mm_sheets.txt"
                    filepath = os.path.join(directory, filename)

                    with open(filepath, 'w', encoding='utf-8') as f:
                        f.write(f"Sheet Layout for {material} {thickness}mm\n")
                        f.write("=" * 60 + "\n\n")

                        for sheet_idx, sheet in enumerate(sheets):
                            f.write(f"Sheet {sheet_idx + 1}\n")
                            f.write("-" * 40 + "\n")

                            # Calculate efficiency
                            sheet_area = self.calculator.sheet_width * self.calculator.sheet_height
                            used_area = sum(rect.width * rect.height for rect in sheet)
                            efficiency = (used_area / sheet_area) * 100

                            f.write(f"Efficiency: {efficiency:.1f}%\n")
                            f.write(f"Parts: {len(sheet)}\n\n")

                            # List parts
                            for rect in sheet:
                                f.write(
                                    f"  {rect.id}: {rect.width:.0f} × {rect.height:.0f} mm "
                                    f"at ({rect.x:.0f}, {rect.y:.0f})\n"
                                )

                            f.write("\n")

                messagebox.showinfo(
                    "Export Complete",
                    f"Sheet layouts exported to {directory}"
                )

            except Exception as e:
                messagebox.showerror(
                    "Export Error",
                    f"Failed to export sheets: {str(e)}"
                )

    def _create_buttons(self, parent: ttk.Frame) -> None:
        """Create dialog buttons."""
        ttk.Button(
            parent,
            text="Print",
            command=self._print_cut_list
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            parent,
            text="Close",
            command=self.dialog.destroy
        ).pack(side=tk.RIGHT, padx=5)

    def _print_cut_list(self) -> None:
        """Print cut list."""
        # TODO: Implement printing functionality
        messagebox.showinfo(
            "Print",
            "Printing functionality will be implemented in a future update."
        )
